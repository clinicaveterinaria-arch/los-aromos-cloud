from openpyxl import load_workbook
from io import BytesIO
import csv
import io

from .vademecum_utils import (
    clean_text,
    title_name,
    normalize_species,
    normalize_route,
    normalize_category,
    normalize_laboratory,
    split_active_ingredients,
)


def pick(row, names):
    normalized = {
        clean_text(k).lower(): clean_text(v)
        for k, v in row.items()
    }

    for name in names:
        key = clean_text(name).lower()
        if key in normalized:
            return normalized[key]

    return ""


def read_rows_from_upload(filename, content):
    filename = (filename or "").lower()

    rows = []

    if filename.endswith(".xlsx"):
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active

        headers = [
            clean_text(cell.value)
            for cell in ws[1]
        ]

        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

        return rows

    text_content = content.decode("utf-8-sig", errors="replace")
    sample = text_content[:1000]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)

    return list(reader)


def parse_vademecum_rows(rows):
    parsed = []
    errors = []

    for index, row in enumerate(rows, start=2):
        active_raw = pick(row, [
            "Principio activo",
            "Activo",
            "Droga",
            "Monodroga",
            "Composición",
            "Composicion",
            "Active ingredient"
        ])

        brand_name = pick(row, [
            "Nombre comercial",
            "Marca",
            "Producto",
            "Especialidad",
            "Commercial name"
        ])

        laboratory = pick(row, [
            "Laboratorio",
            "Elaborador",
            "Titular",
            "Empresa"
        ])

        presentation = pick(row, [
            "Presentación",
            "Presentacion",
            "Forma farmacéutica",
            "Forma farmaceutica",
            "Envase"
        ])

        concentration = pick(row, [
            "Concentración",
            "Concentracion",
            "Composición declarada",
            "Composicion declarada"
        ])

        species = pick(row, [
            "Especie",
            "Especies",
            "Destino",
            "Animales"
        ])

        category = pick(row, [
            "Categoría",
            "Categoria",
            "Rubro",
            "Clase",
            "Grupo terapéutico",
            "Grupo terapeutico"
        ])

        route = pick(row, [
            "Vía",
            "Via",
            "Administración",
            "Administracion"
        ])

        indications = pick(row, [
            "Indicaciones",
            "Uso",
            "Usos",
            "Acción terapéutica",
            "Accion terapeutica"
        ])

        if not active_raw and not brand_name:
            errors.append({
                "row": index,
                "error": "Fila sin principio activo ni nombre comercial."
            })
            continue

        active_names = split_active_ingredients(active_raw)

        if not active_names and brand_name:
            active_names = ["Sin principio activo"]

        for active_name in active_names:
            parsed.append({
                "active_name": title_name(active_name),
                "brand_name": title_name(brand_name),
                "laboratory": normalize_laboratory(laboratory),
                "presentation": clean_text(presentation),
                "concentration": clean_text(concentration),
                "species": normalize_species(species),
                "category": normalize_category(category),
                "route": normalize_route(route),
                "indications": clean_text(indications),
                "source_row": index,
            })

    return parsed, errors
